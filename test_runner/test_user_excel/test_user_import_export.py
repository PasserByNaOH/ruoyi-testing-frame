"""
test_user_import_export.py —— 用户导入导出业务流程测试

流程：创建 Excel → 导入 → DB 验证 → 导出 → 结构校验 + 内容比对
"""

import os
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.apiutil import ApiEngine
from utils.assertions import run_db_verify
from utils.excel_utils import create_user_import_excel
from utils.readyaml import get_testcase_yaml, FILE_PATH
from utils.recordlog import logs


_cases = get_testcase_yaml(
    os.path.join(FILE_PATH["YAML"], "ruoyi", "system", "user_import_export.yaml")
)

# 导入和导出共有的字段（用于内容比对）
_COMMON_FIELDS = ["登录名称", "用户名称", "用户邮箱", "手机号码", "用户性别", "账号状态"]

# 导出文件存放目录
_EXPORT_DIR = os.path.join(os.path.dirname(FILE_PATH["RUNTIME"]), "excel")


@pytest.mark.parametrize(
    "base_info, case",
    _cases,
    ids=[c[1]["case_name"] for c in _cases],
)
def test_import_export(base_url, db_connection, redis_client, base_info, case):
    engine = ApiEngine()
    import_data = case["import"]
    export_data = case["export"]

    # ═══════════════════════════════════════════════════════════
    # 阶段一：导入
    # ═══════════════════════════════════════════════════════════

    # 1. 生成导入 Excel 文件
    excel_path = create_user_import_excel(import_data["rows"], "import_input.xlsx")

    with open(excel_path, "rb") as f:
        excel_bytes = f.read()

    # 2. 调导入 API
    import_case = {
        "case_name": case["case_name"] + " - 导入",
        "url": base_info["import_url"],
        "headers": {"Accept": "application/json, text/plain, */*"},  # 不设 Content-Type，让 requests 自动设 multipart
        "params": {"updateSupport": str(import_data.get("updateSupport", "false")).lower()},
        "files": {
            "file": (
                "users.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        "validations": import_data["validations"],
    }
    engine.specification_yaml(dict(base_info), import_case)

    # 3. DB 验证
    if "db_verify" in import_data:
        run_db_verify(db_connection, engine.replace_load(import_data["db_verify"]))

    # ═══════════════════════════════════════════════════════════
    # 阶段二：导出
    # ═══════════════════════════════════════════════════════════

    export_case = {
        "case_name": case["case_name"] + " - 导出",
        "url": base_info["export_url"],
        "params": export_data["params"],
        "validations": export_data["validations"],
    }
    resp = engine.specification_export(dict(base_info), export_case)

    # 保存导出文件到 data/excel/
    os.makedirs(_EXPORT_DIR, exist_ok=True)
    export_path = os.path.join(_EXPORT_DIR, "export_result.xlsx")
    with open(export_path, "wb") as f:
        f.write(resp.content)
    logs.info(f"导出文件已保存: {export_path}")

    # ═══════════════════════════════════════════════════════════
    # 阶段三：内容比对（导入数据 vs 导出数据）
    # ═══════════════════════════════════════════════════════════

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    export_headers = [cell.value for cell in ws[1]]
    export_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            export_rows.append(dict(zip(export_headers, row)))

    # 行数比对
    assert len(export_rows) == len(import_data["rows"]), (
        f"导出行数 {len(export_rows)} != 导入行数 {len(import_data['rows'])}"
    )

    # 逐行比对共有字段
    for import_user in import_data["rows"]:
        matched = False
        for exp_row in export_rows:
            if exp_row.get("登录名称") == import_user["登录名称"]:
                matched = True
                # 比对 6 个共有字段
                for field in _COMMON_FIELDS:
                    assert str(exp_row.get(field)) == str(import_user[field]), (
                        f"导出值不匹配 [{import_user['登录名称']}]:\n"
                        f"  {field}: export={exp_row.get(field)} vs import={import_user[field]}"
                    )
                # 部门编号 → 部门名称转换比对
                dept_id = import_user["部门编号"]
                dept_rows = db_connection.query(
                    "SELECT dept_name FROM sys_dept WHERE dept_id = %s", [dept_id]
                )
                if dept_rows:
                    assert exp_row.get("部门名称") == dept_rows[0]["dept_name"], (
                        f"部门名称不匹配 [{import_user['登录名称']}]:\n"
                        f"  export={exp_row.get('部门名称')} vs db(dept_id={dept_id})={dept_rows[0]['dept_name']}"
                    )
                break

        assert matched, (
            f"导入用户 {import_user['登录名称']} 未在导出结果中找到\n"
            f"  导出用户列表: {[r.get('登录名称') for r in export_rows]}"
        )
