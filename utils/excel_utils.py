"""
Excel 工具函数：创建导入用 .xlsx、储存导出文件。

所有文件落到 data/excel/ 目录，方便人工检查调试。
"""

import os
from conf.setting import FILE_PATH
from openpyxl import Workbook


# RuoYi 用户导入 Excel 列头（按 SysUser @Excel 注解中非 Type.EXPORT 的字段排序）
_IMPORT_HEADERS = [
    "部门编号",
    "登录名称",
    "用户名称",
    "用户邮箱",
    "手机号码",
    "用户性别",
    "账号状态",
]

# 导出/导入文件存放目录
_EXPORT_DIR = os.path.join(os.path.dirname(FILE_PATH["RUNTIME"]), "excel")


def create_user_import_excel(rows, filename):
    """
    生成用户导入用的 .xlsx 文件，保存到 data/excel/ 目录。

    参数:
        rows:     list[dict] — 每行一个用户，key=列头名，value=单元格值
        filename: str        — 文件名（不含路径），如 "import_users.xlsx"

    返回:
        str — 生成文件的完整绝对路径
    """
    os.makedirs(_EXPORT_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active

    # 写表头
    for col_idx, header in enumerate(_IMPORT_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写数据行
    for row_idx, user in enumerate(rows, start=2):
        for col_idx, header in enumerate(_IMPORT_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=user.get(header, ""))

    filepath = os.path.join(_EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath
