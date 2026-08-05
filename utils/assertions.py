from utils.recordlog import logs

# ═══════════════════════════════════════════════════════════
# HTTP 断言
# ═══════════════════════════════════════════════════════════

def assert_status_code(resp, rule, **kwargs):
     """断言 HTTP 状态码。"""
     assert resp.status_code == rule["expected"], (
          f"HTTP 状态码断言失败\n"
          f"  预期: {rule['expected']}\n"
          f"  实际: {resp.status_code}"
     )

def assert_body_code(resp, rule, **kwargs):
     """断言若依业务状态码（body.code）。"""
     actual = resp.json()
     assert actual.get("code") == rule["expected"], (
          f"body_code 断言失败\n"
          f"  预期 code: {rule['expected']}\n"
          f"  实际响应:   {actual}"
     )


def assert_body_contains(resp, rule, **kwargs):
     """断言响应体包含指定关键字（查整个 JSON 字符串）。"""
     body_text = resp.text
     assert rule["keyword"] in body_text, (
          f"body_contains 断言失败\n"
          f"  预期响应包含: {rule['keyword']}\n"
          f"  实际响应:     {body_text}"
     )


def assert_body_not_contains(resp, rule, **kwargs):
     """断言响应体不包含指定关键字（查整个 JSON 字符串）。"""
     body_text = resp.text
     assert rule["keyword"] not in body_text, (
          f"body_not_contains 断言失败\n"
          f"  预期响应不包含: {rule['keyword']}\n"
          f"  实际响应:       {body_text}"
     )


def assert_token_not_empty(resp, rule, **kwargs):
     """断言响应中包含非空 token。"""
     token = resp.json().get("token", "")
     assert token != "", "登录成功但未返回 token"


def assert_token_absent(resp, rule, **kwargs):
     """断言响应中不包含 token（失败登录场景）。"""
     assert "token" not in resp.json(), (
          f"失败登录不应返回 token，实际返回: {resp.json().get('token')}"
     )

#   验证导出的二进制文件内容是否和数据库的数据相同
def assert_excel_content(resp, rule, **kwargs):
    """
    断言导出 Excel 的内容。

    从 resp.content 解析 .xlsx，提取表头行 + 数据行，
    然后按 rule 中的条件做验证。

    支持的 rule 参数:
        has_headers:  list[str] — 验证 Excel 表头包含指定列名
        row_contains: dict     — 验证至少有一行包含指定 {列名: 值} 组合
        min_rows:     int      — 验证数据行数 >= N
        max_rows:     int      — 验证数据行数 <= N
        row_count:    int      — 验证数据行数 == N
    """
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active

    # 第一行 = 表头
    headers = [cell.value for cell in ws[1]]

    # 后续行 = 数据行（跳过全空行）
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            data_rows.append(dict(zip(headers, row)))

    # ── 断言 ──
    if "has_headers" in rule:
        for h in rule["has_headers"]:
            assert h in headers, (
                f"Excel 缺少列头: {h}\n"
                f"  实际列头: {headers}"
            )

    if "row_contains" in rule:
        expected = rule["row_contains"]
        found = False
        for row in data_rows:
            if all(str(row.get(k)) == str(v) for k, v in expected.items()):
                found = True
                break
        assert found, (
            f"Excel 中未找到匹配行: {expected}\n"
            f"  实际数据({len(data_rows)}行): {data_rows}"
        )

    if "min_rows" in rule:
        assert len(data_rows) >= rule["min_rows"], (
            f"Excel 数据行数 {len(data_rows)} < 预期最少 {rule['min_rows']}"
        )

    if "max_rows" in rule:
        assert len(data_rows) <= rule["max_rows"], (
            f"Excel 数据行数 {len(data_rows)} > 预期最多 {rule['max_rows']}"
        )

    if "row_count" in rule:
        assert len(data_rows) == rule["row_count"], (
            f"Excel 数据行数 {len(data_rows)} != 预期 {rule['row_count']}"
        )


# ═══════════════════════════════════════════════════════════
# DataScope 断言（依赖 db）
# ═══════════════════════════════════════════════════════════

def assert_rows_in_scope(resp, rule, db=None, **kwargs):
    """
    断言响应 body.rows 中每一行都在指定用户的 DataScope 范围内。

    规则参数:
        username:   str  — 以该用户的 DataScope 为基准
        dept_field: str  — rows 中部门 ID 的字段名（默认 "deptId"）
        user_field: str  — rows 中用户 ID 的字段名（默认 "userId"）
    """
    assert db is not None, "rows_in_scope 需要 db 参数（specification_yaml 传入）"

    body = resp.json()
    rows = body.get("rows", [])
    username = rule["username"]
    dept_field = rule.get("dept_field", "deptId")
    user_field = rule.get("user_field", "userId")

    # ── 复用 DataScopeAspect 逻辑：计算用户的允许部门集合 ──
    user_recs = db.query(
        "SELECT user_id, dept_id FROM sys_user WHERE user_name = %s", [username])
    assert user_recs, f"用户不存在: {username}"
    user_id = user_recs[0]["user_id"]
    dept_id = user_recs[0]["dept_id"]

    roles = db.query(
        "SELECT r.role_id, r.data_scope FROM sys_role r "
        "JOIN sys_user_role ur ON r.role_id = ur.role_id "
        "WHERE ur.user_id = %s AND r.status = '0' AND r.del_flag = '0'", [user_id])

    allowed_dept_ids = set()
    is_self_only = False

    for role in roles:
        ds = role["data_scope"]
        if ds == "1":          # 全部数据权限 → 不限
            return
        elif ds == "2":         # 自定义 → sys_role_dept 查关联部门
            depts = db.query(
                "SELECT dept_id FROM sys_role_dept WHERE role_id = %s",
                [role["role_id"]])
            allowed_dept_ids.update(d["dept_id"] for d in depts)
        elif ds == "3":         # 本部门
            allowed_dept_ids.add(dept_id)
        elif ds == "4":         # 本部门及以下
            subs = db.query(
                "SELECT dept_id FROM sys_dept "
                "WHERE dept_id = %s OR FIND_IN_SET(%s, ancestors)",
                [dept_id, dept_id])
            allowed_dept_ids.update(d["dept_id"] for d in subs)
        elif ds == "5":         # 仅本人
            is_self_only = True

    # ── 逐行验证 ──
    if is_self_only and not allowed_dept_ids:
        assert len(rows) <= 1, (
            f"rows_in_scope 失败 [{username}]：仅本人权限但返回了 {len(rows)} 条\n"
            f"  userIds: {[r.get(user_field) for r in rows]}"
        )
        if rows:
            assert rows[0].get(user_field) == user_id, (
                f"rows_in_scope 失败 [{username}]：仅本人权限但返回了其他用户\n"
                f"  预期 userId={user_id}\n"
                f"  实际 userId={rows[0].get(user_field)}"
            )

    elif allowed_dept_ids:
        for row in rows:
            actual_dept = row.get(dept_field)
            assert actual_dept in allowed_dept_ids, (
                f"rows_in_scope 失败 [{username}]："
                f"{dept_field}={actual_dept} 不在 {sorted(allowed_dept_ids)} 内"
            )


# type 字符串 → 断言函数的映射表
VALIDATORS = {
     "status_code":       assert_status_code,
     "body_code":         assert_body_code,
     "body_contains":     assert_body_contains,
     "body_not_contains": assert_body_not_contains,
     "token_not_empty":   assert_token_not_empty,
     "token_absent":      assert_token_absent,
     "excel_content":     assert_excel_content,
     "rows_in_scope":     assert_rows_in_scope,
}

def run_validations(resp, validations, **kwargs):
     """
     遍历验证规则列表，逐个执行断言。
     加新断言类型：写一个 assert_xxx 函数 → 在 VALIDATORS 里加一行即可。

     kwargs 透传给断言函数（例如 db=xxx 给 rows_in_scope）。
     """
     for rule in validations:
          validate_type = rule["type"]
          validate_func = VALIDATORS.get(validate_type)
          if validate_func is None:
               raise ValueError(f"不支持的断言类型: {validate_type}")
          validate_func(resp, rule, **kwargs)
          logs.info(f"断言通过: {validate_type}")


# ═══════════════════════════════════════════════════════════
# DB 验证
# ═══════════════════════════════════════════════════════════

def coerce_db_param(value):
    """将 replace_load 反序列化后的字符串还原为适合 SQL 参数的类型。

    问题背景：YAML 中的 `${get_runtime(created_user_id)}` 经过 replace_load 后，
    数字 191 变成了字符串 "191"（因为 str(result) 拼接）。
    MySQL 的 WHERE user_id = '191' 字符串与 bigint 列可能不匹配，
    需要转回 int。
    """
    if not isinstance(value, str):
        return value
    # int：纯数字或负号开头
    if value.isdigit() or (value.startswith("-") and value[1:].isdigit()):
        return int(value)
    # bool
    if value.lower() in ("true", "false"):
        return value.lower() == "true"
    return value


def run_db_verify(db, rules):
    """遍历 db_verify 规则列表，对 MySQL 做声明式断言。

    db 是 ConnectMysql 实例（autocommit=True，每次 query 都是最新快照）。

    支持 4 种 expect 类型：
      exists     — SELECT COUNT(*) >= 1
      count      — SELECT COUNT(*) == value
      eq         — SELECT column == value
      not_empty  — SELECT column IS NOT NULL AND != ''
    """
    for rule in rules:
        expect = rule["expect"]
        table = rule["table"]
        where = rule["where"]

        pairs = [f"{k} = %s" for k in where]
        clause = " AND ".join(pairs)
        params = [coerce_db_param(v) for v in where.values()]

        if expect == "exists":
            sql = f"SELECT COUNT(*) FROM {table} WHERE {clause}"
            actual = db.query(sql, params)[0]["COUNT(*)"]
            assert actual >= 1, (
                f"DB验证失败 [{rule['desc']}]\n"
                f"  SQL: {sql}\n  params: {params}\n"
                f"  COUNT(*): {actual} (预期 >=1)"
            )

        elif expect == "count":
            sql = f"SELECT COUNT(*) FROM {table} WHERE {clause}"
            actual = db.query(sql, params)[0]["COUNT(*)"]
            assert actual == rule["value"], (
                f"DB验证失败 [{rule['desc']}]\n"
                f"  SQL: {sql}\n  params: {params}\n"
                f"  COUNT(*): {actual} (预期 {rule['value']})"
            )

        elif expect == "eq":
            sql = f"SELECT {rule['column']} FROM {table} WHERE {clause}"
            rows = db.query(sql, params)
            assert rows, (
                f"DB验证失败 [{rule['desc']}]\n"
                f"  SQL: {sql}\n  params: {params}\n"
                f"  查询结果: 0 行，用户可能不存在"
            )
            actual = rows[0][rule["column"]]
            assert str(actual) == str(rule["value"]), (
                f"DB验证失败 [{rule['desc']}]\n"
                f"  SQL: {sql}\n  params: {params}\n"
                f"  实际: {actual!r}\n  预期: {rule['value']!r}"
            )

        elif expect == "not_empty":
            sql = f"SELECT {rule['column']} FROM {table} WHERE {clause}"
            rows = db.query(sql, params)
            assert rows, (
                f"DB验证失败 [{rule['desc']}]\n"
                f"  SQL: {sql}\n  params: {params}\n"
                f"  查询结果: 0 行，用户可能不存在"
            )
            actual = rows[0][rule["column"]]
            assert actual not in (None, ""), (
                f"DB验证失败 [{rule['desc']}]\n"
                f"  SQL: {sql}\n  params: {params}\n"
                f"  实际: {actual!r}（预期非空）"
            )

        else:
            raise ValueError(f"不支持的 DB 验证类型: {expect}")

        logs.info(f"DB验证通过: {rule['desc']}")


# ═══════════════════════════════════════════════════════════
# 自测入口：用假响应验证所有断言类型
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    class MockResponse:
        """模拟 requests.Response 对象。"""
        def __init__(self, status_code, json_data):
            self.status_code = status_code
            self._json = json_data

        def json(self):
            return self._json

    # 模拟登录成功响应
    success_resp = MockResponse(200, {
        "code": 200,
        "msg": "操作成功",
        "token": "eyJhbGciOi..."
    })

    # 模拟登录失败响应
    fail_resp = MockResponse(200, {
        "code": 500,
        "msg": "验证码错误"
    })

    print("--- 测试 1: status_code 断言（成功） ---")
    run_validations(success_resp, [{"type": "status_code", "expected": 200}])

    print("--- 测试 2: body_code 断言（成功 + 失败） ---")
    run_validations(success_resp, [{"type": "body_code", "expected": 200}])

    try:
        run_validations(fail_resp, [{"type": "body_code", "expected": 200}])
        print("  预期应失败但通过了！")
    except AssertionError as e:
        print(f"  正确捕获到失败: body_code 500 != 200")

    print("--- 测试 3: body_contains 断言 ---")
    run_validations(success_resp, [{"type": "body_contains", "keyword": "操作成功"}])

    print("--- 测试 4: token_not_empty 断言 ---")
    run_validations(success_resp, [{"type": "token_not_empty"}])

    print("--- 测试 5: token_absent 断言 ---")
    run_validations(fail_resp, [{"type": "token_absent"}])

    print("--- 测试 6: 不支持的 type 应报错 ---")
    try:
        run_validations(success_resp, [{"type": "unknown_type"}])
    except ValueError as e:
        print(f"  正确捕获: {e}")

    print("\n所有断言类型验证完毕！")